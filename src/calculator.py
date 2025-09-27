import pandas as pd
import numpy as np
import json
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook

from load_data import lire_feuille_wide, extraire_poids  # import direct

def extraire_toutes_categories(d):
    """Extrait récursivement toutes les clés terminales d'un dict JSON hiérarchique."""
    result = set()
    if isinstance(d, dict):
        for k, v in d.items():
            result.add(k)
            result |= extraire_toutes_categories(v)
    elif isinstance(d, list):
        for v in d:
            result |= extraire_toutes_categories(v)
    return result

def calculer_ipc(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule l'IPC global d'un panier en utilisant les poids de config/weights.json
    et insère/réécrit les résultats dans une colonne fixe 'IPC (%)'.
    """

    # --- Charger la feuille en wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- S'assurer que la colonne 'date' est bien au format YYYY-MM
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"]).dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index).to_period("M")

    # --- Transformer les arguments en période mensuelle
    date_debut = pd.Period(date_debut, freq="M")
    date_fin = pd.Period(date_fin, freq="M")

    # --- Filtrer la période
    df = df.loc[date_debut:date_fin]

    # --- Charger les poids
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)

    poids_feuille = extraire_poids(all_weights.get(feuille, {}))

    if not poids_feuille:
        raise ValueError(f"Aucun poids trouvé pour la feuille {feuille} dans weights.json")

    colonnes_valides = [col for col in df.columns if col in poids_feuille]
    if not colonnes_valides:
        raise ValueError("Aucune correspondance entre colonnes du fichier Excel et weights.json")

    # --- Calcul de l’IPC (moyenne pondérée)
    numerateur = sum(df[col] * poids_feuille[col] for col in colonnes_valides)
    denominateur = sum(poids_feuille[col] for col in colonnes_valides)

    df["IPC (%)"] = (numerateur / denominateur).round(2)

    # --- Insérer dans Excel avec openpyxl en réécrivant toujours dans 'IPC (%)'
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    # Trouver ou créer la colonne "IPC (%)"
    header_row = 1
    col_index_ipc = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "IPC (%)":
            col_index_ipc = col
            break

    if col_index_ipc is None:
        col_index_ipc = ws.max_column + 1
        ws.cell(row=header_row, column=col_index_ipc, value="IPC (%)")

    # Construire un dictionnaire {periode: valeur}
    ipc_dict = df["IPC (%)"].to_dict()

    # Écrire les valeurs au bon endroit
    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value  # on suppose la date est en première colonne
        if cell_date is None:
            continue

        try:
            periode = pd.to_datetime(cell_date).to_period("M")
        except Exception:
            continue

        if periode in ipc_dict:
            ws.cell(row=row, column=col_index_ipc, value=float(ipc_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df

def calculer_ipc_core_noncore(nom_fichier: str, feuille_core: str, feuille_non_core: str,
                              date_debut: str, date_fin: str):
    """
    Calcule l'IPC core et l'IPC non-core et les insère dans les colonnes
    'IPC Core (%)' et 'IPC Non Core (%)' des feuilles correspondantes.
    Si la colonne existe déjà, elle est réécrite (pas de nouvelle colonne ajoutée).
    """

    # --- Charger config/weights.json ---
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)

    # --------------------------------------------------
    # Fonction interne pour calculer et insérer un IPC
    # --------------------------------------------------
    def traiter_feuille(feuille: str, nom_colonne: str):
        # Lire feuille wide
        df = lire_feuille_wide(nom_fichier, feuille)

        # Normaliser les dates
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"]).dt.to_period("M")
            df.set_index("date", inplace=True)
        else:
            df.index = pd.to_datetime(df.index).to_period("M")

        # Filtrer période
        d_debut = pd.Period(date_debut, freq="M")
        d_fin = pd.Period(date_fin, freq="M")
        df = df.loc[d_debut:d_fin]

        # Extraire les poids
        poids_feuille = extraire_poids(all_weights.get(feuille, {}))
        if not poids_feuille:
            raise ValueError(f"Aucun poids trouvé pour la feuille {feuille} dans weights.json")

        # Colonnes valides
        colonnes_valides = [col for col in df.columns if col in poids_feuille]
        if not colonnes_valides:
            raise ValueError(f"Aucune correspondance entre colonnes Excel et poids pour {feuille}")

        # Calcul IPC pondéré
        numerateur = sum(df[col] * poids_feuille[col] for col in colonnes_valides)
        denominateur = sum(poids_feuille[col] for col in colonnes_valides)
        df[nom_colonne] = (numerateur / denominateur).round(2)

        # Insérer dans Excel
        wb = load_workbook(nom_fichier)
        ws = wb[feuille]

        # Vérifier si la colonne existe déjà
        header_row = 1
        col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col).value == nom_colonne:
                col_index = col
                break

        # Si pas trouvée → créer à la fin
        if col_index is None:
            col_index = ws.max_column + 1
            ws.cell(row=header_row, column=col_index, value=nom_colonne)

        # Construire dict {periode: valeur}
        ipc_dict = df[nom_colonne].to_dict()

        # Écrire dans la bonne colonne
        for row in range(2, ws.max_row + 1):
            cell_date = ws.cell(row=row, column=1).value
            if cell_date is None:
                continue
            try:
                periode = pd.to_datetime(cell_date).to_period("M")
            except Exception:
                continue
            if periode in ipc_dict:
                ws.cell(row=row, column=col_index, value=float(ipc_dict[periode]))

        wb.save(nom_fichier)
        wb.close()

        return df

    # --- Appliquer aux deux feuilles ---
    df_core = traiter_feuille(feuille_core, "IPC Core (%)")
    df_non_core = traiter_feuille(feuille_non_core, "IPC Non Core (%)")

    return df_core, df_non_core

def calculer_inflation_mom(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule l'inflation en glissement mensuel (mom) à partir des valeurs d'IPC d'une feuille
    et insère/réécrit les résultats dans une colonne fixe 'Inflation (%, mom)'.
    """

    # --- Charger la feuille en wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- Assurer que la colonne 'date' est bien au format YYYY-MM
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"]).dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index).to_period("M")

    # --- Transformer les arguments en période mensuelle
    date_debut = pd.Period(date_debut, freq="M")
    date_fin = pd.Period(date_fin, freq="M")

    # --- Filtrer la période
    df = df.loc[date_debut:date_fin]

    # --- Chercher la colonne IPC de référence
    for col in ["IPC (%)", "IPC Core (%)", "IPC Non Core (%)"]:
        if col in df.columns:
            col_ipc = col
            break
    else:
        raise ValueError("Aucune colonne IPC trouvée (IPC (%), IPC Core (%), ou IPC Non Core (%)).")

    # --- Calcul de l’inflation mom : (IPC_t / IPC_t-1 - 1) * 100
    df["Inflation (%, mom)"] = ((df[col_ipc] / df[col_ipc].shift(1) - 1) * 100).round(2)

    # --- Insérer dans Excel avec openpyxl
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    # Trouver ou créer la colonne "Inflation (%, mom)"
    header_row = 1
    col_index_inflation = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Inflation (%, mom)":
            col_index_inflation = col
            break

    if col_index_inflation is None:
        col_index_inflation = ws.max_column + 1
        ws.cell(row=header_row, column=col_index_inflation, value="Inflation (%, mom)")

    # Construire un dictionnaire {periode: valeur}
    infl_dict = df["Inflation (%, mom)"].to_dict()

    # Écrire les valeurs au bon endroit
    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value  # date supposée en première colonne
        if cell_date is None:
            continue

        try:
            periode = pd.to_datetime(cell_date).to_period("M")
        except Exception:
            continue

        if periode in infl_dict and pd.notna(infl_dict[periode]):
            ws.cell(row=row, column=col_index_inflation, value=float(infl_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df


def calculer_inflation_yoy(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule l'inflation en glissement annuel (yoy) à partir des valeurs d'IPC d'une feuille
    et insère/réécrit les résultats dans une colonne fixe 'Inflation (%, yoy)'.
    """

    # --- Charger la feuille en wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- Assurer que la colonne 'date' est bien au format YYYY-MM
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"]).dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index).to_period("M")

    # --- Transformer les arguments en période mensuelle
    date_debut = pd.Period(date_debut, freq="M")
    date_fin = pd.Period(date_fin, freq="M")

    # --- Filtrer la période
    df = df.loc[date_debut:date_fin]

    # --- Chercher la colonne IPC de référence
    for col in ["IPC (%)", "IPC Core (%)", "IPC Non Core (%)"]:
        if col in df.columns:
            col_ipc = col
            break
    else:
        raise ValueError("Aucune colonne IPC trouvée (IPC (%), IPC Core (%), ou IPC Non Core (%)).")

    # --- Calcul de l’inflation yoy : (IPC_t / IPC_t-12 - 1) * 100
    df["Inflation (%, yoy)"] = ((df[col_ipc] / df[col_ipc].shift(12) - 1) * 100).round(2)

    # --- Insérer dans Excel avec openpyxl
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    # Trouver ou créer la colonne "Inflation (%, yoy)"
    header_row = 1
    col_index_inflation = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Inflation (%, yoy)":
            col_index_inflation = col
            break

    if col_index_inflation is None:
        col_index_inflation = ws.max_column + 1
        ws.cell(row=header_row, column=col_index_inflation, value="Inflation (%, yoy)")

    # Construire un dictionnaire {periode: valeur}
    infl_dict = df["Inflation (%, yoy)"].to_dict()

    # Écrire les valeurs au bon endroit
    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value  # date supposée en première colonne
        if cell_date is None:
            continue

        try:
            periode = pd.to_datetime(cell_date).to_period("M")
        except Exception:
            continue

        if periode in infl_dict and pd.notna(infl_dict[periode]):
            ws.cell(row=row, column=col_index_inflation, value=float(infl_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df

def calculer_inflation_elements_mom(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule l'inflation mensuelle (MoM, %) uniquement pour les colonnes
    définies dans categories.json + weights.json, et insère les résultats
    dans la feuille Excel (Inflation_<élément>_MoM (%)).
    """

    # --- Charger la feuille wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- Normaliser date -> Period M
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index, errors="coerce").to_period("M")

    # --- Périodes demandées
    d_debut = pd.Period(date_debut, freq="M")
    d_fin = pd.Period(date_fin, freq="M")
    df = df.loc[d_debut:d_fin].copy()

    # --- Charger poids et catégories
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"
    CATEG_PATH = BASE_DIR / "config" / "categories.json"

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)
    with open(CATEG_PATH, "r", encoding="utf-8") as f:
        categories = json.load(f)

    poids_feuille = extraire_poids(all_weights.get(feuille, {}))
    if not poids_feuille:
        raise ValueError(f"Aucun poids trouvé pour la feuille '{feuille}' dans weights.json")

    # --- Normalisation des noms pour comparer
    poids_set = {k.strip().lower() for k in poids_feuille.keys()}
    categ_set = {c.strip().lower() for c in extraire_toutes_categories(categories)}
    df_cols_set = {c.strip().lower() for c in df.columns}

    # Colonnes valides = intersection
    colonnes_valides = [
        col for col in df.columns
        if col.strip().lower() in poids_set and col.strip().lower() in categ_set
    ]

    if not colonnes_valides:
        raise ValueError(
            f"Aucune colonne valide trouvée.\n"
            f"Colonnes Excel = {sorted(df.columns.tolist())}\n"
            f"Colonnes weights.json = {sorted(poids_feuille.keys())}\n"
            f"Colonnes categories.json = {sorted(list(categ_set))}"
        )

    # --- Calcul inflation MoM uniquement pour colonnes valides
    df_infl = pd.DataFrame(index=df.index)
    for col in colonnes_valides:
        prev1 = df[col].shift(1)  # <-- MoM = (t - t-1) / t-1
        infl = ((df[col] - prev1) / prev1) * 100
        df_infl[f"Inflation_MoM (%)_{col}"] = infl.replace([np.inf, -np.inf], np.nan).round(2)

    # --- Écriture dans Excel
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    # Dictionnaire colonnes existantes
    header_row = 1
    col_map = {ws.cell(row=header_row, column=c).value: c
               for c in range(1, ws.max_column + 1)}

    for col_name in df_infl.columns:
        # Trouver ou créer la colonne
        if col_name in col_map:
            col_index = col_map[col_name]
        else:
            col_index = ws.max_column + 1
            ws.cell(row=header_row, column=col_index, value=col_name)
            col_map[col_name] = col_index

        # Écrire les valeurs
        infl_dict = df_infl[col_name].to_dict()
        for r in range(2, ws.max_row + 1):
            cell_date = ws.cell(row=r, column=1).value
            if cell_date is None:
                continue
            try:
                periode = pd.to_datetime(cell_date, errors="coerce").to_period("M")
            except Exception:
                continue
            if periode in infl_dict:
                val = infl_dict[periode]
                if pd.notna(val):
                    ws.cell(row=r, column=col_index, value=float(val))

    wb.save(nom_fichier)
    wb.close()

    return df_infl

def calculer_inflation_elements_yoy(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule l'inflation annuelle (YoY, %) uniquement pour les colonnes
    définies dans categories.json + weights.json, et insère les résultats
    dans la feuille Excel (Inflation_<élément> (%)).
    """

    # --- Charger la feuille wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- Normaliser date -> Period M
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index, errors="coerce").to_period("M")

    # --- Périodes demandées
    d_debut = pd.Period(date_debut, freq="M")
    d_fin = pd.Period(date_fin, freq="M")
    df = df.loc[d_debut:d_fin].copy()

    # --- Charger poids et catégories
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"
    CATEG_PATH = BASE_DIR / "config" / "categories.json"

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)
    with open(CATEG_PATH, "r", encoding="utf-8") as f:
        categories = json.load(f)

    poids_feuille = extraire_poids(all_weights.get(feuille, {}))
    if not poids_feuille:
        raise ValueError(f"Aucun poids trouvé pour la feuille '{feuille}' dans weights.json")

    # --- Normalisation des noms pour comparer
    poids_set = {k.strip().lower() for k in poids_feuille.keys()}
    categ_set = {c.strip().lower() for c in extraire_toutes_categories(categories)}
    df_cols_set = {c.strip().lower() for c in df.columns}

    # Colonnes valides = intersection
    colonnes_valides = [
        col for col in df.columns
        if col.strip().lower() in poids_set and col.strip().lower() in categ_set
    ]

    if not colonnes_valides:
        raise ValueError(
            f"Aucune colonne valide trouvée.\n"
            f"Colonnes Excel = {sorted(df.columns.tolist())}\n"
            f"Colonnes weights.json = {sorted(poids_feuille.keys())}\n"
            f"Colonnes categories.json = {sorted(list(categ_set))}"
        )

    # --- Calcul inflation YoY uniquement pour colonnes valides
    df_infl = pd.DataFrame(index=df.index)
    for col in colonnes_valides:
        prev12 = df[col].shift(12)
        infl = ((df[col] - prev12) / prev12) * 100
        df_infl[f"Inflation_YoY (%)_{col}"] = infl.replace([np.inf, -np.inf], np.nan).round(2)

    # --- Écriture dans Excel
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    # Dictionnaire colonnes existantes
    header_row = 1
    col_map = {ws.cell(row=header_row, column=c).value: c
               for c in range(1, ws.max_column + 1)}

    for col_name in df_infl.columns:
        # Trouver ou créer la colonne
        if col_name in col_map:
            col_index = col_map[col_name]
        else:
            col_index = ws.max_column + 1
            ws.cell(row=header_row, column=col_index, value=col_name)
            col_map[col_name] = col_index

        # Écrire les valeurs
        infl_dict = df_infl[col_name].to_dict()
        for r in range(2, ws.max_row + 1):
            cell_date = ws.cell(row=r, column=1).value
            if cell_date is None:
                continue
            try:
                periode = pd.to_datetime(cell_date, errors="coerce").to_period("M")
            except Exception:
                continue
            if periode in infl_dict:
                val = infl_dict[periode]
                if pd.notna(val):
                    ws.cell(row=r, column=col_index, value=float(val))

    wb.save(nom_fichier)
    wb.close()

    return df_infl

def calculer_contributions_pp_mom(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule les contributions mensuelles (MoM, en pp) pour CHAQUE élément du panier
    (selon categories.json) et écrit une colonne par élément dans Excel.

    Retourne :
      df_contrib : DataFrame avec toutes les colonnes Contrib_<élément>_MoM
      ipc_info   : DataFrame avec IPC_level et IPC_mom
    """
    # --- Charger la feuille wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- Normaliser date -> Period M
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index, errors="coerce").to_period("M")

    # --- Périodes demandées
    d_debut = pd.Period(date_debut, freq="M")
    d_fin = pd.Period(date_fin, freq="M")
    df = df.loc[d_debut:d_fin].copy()

    # --- Charger poids et catégories
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"
    CATEG_PATH = BASE_DIR / "config" / "categories.json"

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)
    with open(CATEG_PATH, "r", encoding="utf-8") as f:
        categories = json.load(f)

    poids_feuille = extraire_poids(all_weights.get(feuille, {}))
    if not poids_feuille:
        raise ValueError(f"Aucun poids trouvé pour la feuille '{feuille}' dans weights.json")

    colonnes_valides = [col for col in df.columns if col in poids_feuille]
    if not colonnes_valides:
        raise ValueError("Aucune colonne du fichier Excel ne correspond aux poids du panier.")

    # --- IPC global
    numer = sum(df[col].astype(float) * float(poids_feuille[col]) for col in colonnes_valides)
    denom = sum(float(poids_feuille[col]) for col in colonnes_valides)
    ipc_level = (numer / denom).rename("IPC_level")
    ipc_info = ipc_level.to_frame()
    ipc_info["IPC_prev1"] = ipc_info["IPC_level"].shift(1)
    ipc_info["IPC_mom_pct"] = ((ipc_info["IPC_level"] - ipc_info["IPC_prev1"])
                                / ipc_info["IPC_prev1"]) * 100

    # --- Calcul contributions détaillées (MoM)
    df_contrib = pd.DataFrame(index=df.index)
    for col in colonnes_valides:
        poids_i = float(poids_feuille[col])
        delta = df[col] - df[col].shift(1)  # <-- MoM
        contrib = (delta / ipc_info["IPC_prev1"]) * (poids_i / denom) * 100
        df_contrib[f"Contrib_MoM_{col} (pp)"] = contrib.replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3)

    # --- Écriture Excel (une colonne par élément, ordre de categories.json)
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    header_row = 1
    col_map = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Parcours hiérarchique des catégories pour garder l’ordre
    for cat, elements in categories.items():
        for elem in elements:
            col_name = f"Contrib_MoM_{elem} (pp)"
            if col_name in df_contrib.columns:
                # Trouver ou créer la colonne
                if col_name in col_map:
                    col_index = col_map[col_name]
                else:
                    col_index = ws.max_column + 1
                    ws.cell(row=header_row, column=col_index, value=col_name)
                    col_map[col_name] = col_index

                contrib_dict = df_contrib[col_name].to_dict()

                for r in range(2, ws.max_row + 1):
                    cell_date = ws.cell(row=r, column=1).value
                    if cell_date is None:
                        continue
                    try:
                        periode = pd.to_datetime(cell_date, errors="coerce").to_period("M")
                    except Exception:
                        continue
                    if periode in contrib_dict:
                        ws.cell(row=r, column=col_index, value=float(contrib_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df_contrib, ipc_info

def calculer_contributions_pp_yoy(nom_fichier: str, feuille: str, date_debut: str, date_fin: str):
    """
    Calcule les contributions en pp pour CHAQUE élément du panier
    (selon categories.json) et écrit une colonne par élément dans Excel.

    Retourne :
      df_contrib : DataFrame avec toutes les colonnes Contrib_<élément>
      ipc_info   : DataFrame avec IPC_level et IPC_yoy
    """
    # --- Charger la feuille wide
    df = lire_feuille_wide(nom_fichier, feuille)

    # --- Normaliser date -> Period M
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M")
        df.set_index("date", inplace=True)
    else:
        df.index = pd.to_datetime(df.index, errors="coerce").to_period("M")

    # --- Périodes demandées
    d_debut = pd.Period(date_debut, freq="M")
    d_fin = pd.Period(date_fin, freq="M")
    df = df.loc[d_debut:d_fin].copy()

    # --- Charger poids et catégories
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"
    CATEG_PATH = BASE_DIR / "config" / "categories.json"

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)
    with open(CATEG_PATH, "r", encoding="utf-8") as f:
        categories = json.load(f)

    poids_feuille = extraire_poids(all_weights.get(feuille, {}))
    if not poids_feuille:
        raise ValueError(f"Aucun poids trouvé pour la feuille '{feuille}' dans weights.json")

    colonnes_valides = [col for col in df.columns if col in poids_feuille]
    if not colonnes_valides:
        raise ValueError("Aucune colonne du fichier Excel ne correspond aux poids du panier.")

    # --- IPC global
    numer = sum(df[col].astype(float) * float(poids_feuille[col]) for col in colonnes_valides)
    denom = sum(float(poids_feuille[col]) for col in colonnes_valides)
    ipc_level = (numer / denom).rename("IPC_level")
    ipc_info = ipc_level.to_frame()
    ipc_info["IPC_prev12"] = ipc_info["IPC_level"].shift(12)
    ipc_info["IPC_yoy_pct"] = ((ipc_info["IPC_level"] - ipc_info["IPC_prev12"])
                                / ipc_info["IPC_prev12"]) * 100

    # --- Calcul contributions détaillées
    df_contrib = pd.DataFrame(index=df.index)
    for col in colonnes_valides:
        poids_i = float(poids_feuille[col])
        delta = df[col] - df[col].shift(12)
        contrib = (delta / ipc_info["IPC_prev12"]) * (poids_i / denom) * 100
        df_contrib[f"Contrib_YoY_{col} (pp)"] = contrib.replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3)

    # --- Écriture Excel (une colonne par élément, ordre de categories.json)
    wb = load_workbook(nom_fichier)
    ws = wb[feuille]

    header_row = 1
    col_map = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Parcours hiérarchique des catégories pour garder l’ordre
    for cat, elements in categories.items():
        for elem in elements:
            col_name = f"Contrib_YoY_{elem} (pp)"
            if col_name in df_contrib.columns:
                # Trouver ou créer la colonne
                if col_name in col_map:
                    col_index = col_map[col_name]
                else:
                    col_index = ws.max_column + 1
                    ws.cell(row=header_row, column=col_index, value=col_name)
                    col_map[col_name] = col_index

                contrib_dict = df_contrib[col_name].to_dict()

                for r in range(2, ws.max_row + 1):
                    cell_date = ws.cell(row=r, column=1).value
                    if cell_date is None:
                        continue
                    try:
                        periode = pd.to_datetime(cell_date, errors="coerce").to_period("M")
                    except Exception:
                        continue
                    if periode in contrib_dict:
                        ws.cell(row=r, column=col_index, value=float(contrib_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df_contrib, ipc_info


def calculer_contributions_core_noncore_mom(nom_fichier: str,
                                            feuille_core: str,
                                            feuille_noncore: str,
                                            feuille_categories: str,
                                            date_debut: str,
                                            date_fin: str):
    """
    Calcule la contribution mensuelle (MoM, en pp) du Core et du Non-Core
    dans l'inflation globale (feuille 'categories').
    Insère les colonnes 'Contrib_Core_MoM (pp)' et 'Contrib_Non_Core_MoM (pp)' dans la feuille categories.

    Retourne :
        df_contrib : DataFrame avec Contrib_Core_MoM et Contrib_Non_Core_MoM
        ipc_info   : DataFrame avec IPC_level et IPC_mom_pct
    """

    # --- 1. Charger les 3 feuilles
    df_core = lire_feuille_wide(nom_fichier, feuille_core)
    df_noncore = lire_feuille_wide(nom_fichier, feuille_noncore)
    df_cat = lire_feuille_wide(nom_fichier, feuille_categories)

    # --- 2. Normaliser les dates
    for df in (df_core, df_noncore, df_cat):
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M")
            df.set_index("date", inplace=True)
        else:
            df.index = pd.to_datetime(df.index, errors="coerce").to_period("M")

    # --- 3. Restreindre à la période demandée
    d_debut = pd.Period(date_debut, freq="M")
    d_fin = pd.Period(date_fin, freq="M")
    df_core = df_core.loc[d_debut:d_fin].copy()
    df_noncore = df_noncore.loc[d_debut:d_fin].copy()
    df_cat = df_cat.loc[d_debut:d_fin].copy()

    # --- 4. Charger les poids
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)

    poids_core = extraire_poids(all_weights.get(feuille_core, {}))
    poids_noncore = extraire_poids(all_weights.get(feuille_noncore, {}))
    poids_cat = extraire_poids(all_weights.get(feuille_categories, {}))

    # --- 5. Colonnes valides
    colonnes_core = [c for c in df_core.columns if c in poids_core]
    colonnes_noncore = [c for c in df_noncore.columns if c in poids_noncore]
    colonnes_cat = [c for c in df_cat.columns if c in poids_cat]

    if not colonnes_core or not colonnes_noncore or not colonnes_cat:
        raise ValueError("Colonnes manquantes ou incohérence entre Excel et weights.json")

    # --- 6. IPC global
    numer_cat = sum(df_cat[col] * poids_cat[col] for col in colonnes_cat)
    denom_cat = sum(poids_cat[col] for col in colonnes_cat)
    ipc_level = (numer_cat / denom_cat).rename("IPC_level")
    ipc_prev1 = ipc_level.shift(1)

    # --- 7. IPC Core et Non-Core
    numer_core = sum(df_core[col] * poids_core[col] for col in colonnes_core)
    denom_core = sum(poids_core[col] for col in colonnes_core)
    ipc_core = numer_core / denom_core

    numer_noncore = sum(df_noncore[col] * poids_noncore[col] for col in colonnes_noncore)
    denom_noncore = sum(poids_noncore[col] for col in colonnes_noncore)
    ipc_noncore = numer_noncore / denom_noncore

    # --- 8. Contributions MoM (pp)
    contrib_core = ((ipc_core - ipc_core.shift(1)) / ipc_prev1) * (denom_core / denom_cat) * 100
    contrib_noncore = ((ipc_noncore - ipc_noncore.shift(1)) / ipc_prev1) * (denom_noncore / denom_cat) * 100

    df_contrib = pd.DataFrame({
        "Contrib_Core_MoM (pp)": contrib_core.replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3),
        "Contrib_Non_Core_MoM (pp)": contrib_noncore.replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3),
    })

    # --- 9. IPC global MoM %
    ipc_mom_pct = ((ipc_level - ipc_prev1) / ipc_prev1) * 100
    ipc_info = pd.DataFrame({
        "IPC_level": ipc_level,
        "IPC_mom_pct": ipc_mom_pct
    })

    # --- 10. Écriture dans Excel
    wb = load_workbook(nom_fichier)
    ws = wb[feuille_categories]
    header_row = 1
    col_map = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}

    for col_name in df_contrib.columns:
        if col_name in col_map:
            col_index = col_map[col_name]
        else:
            col_index = ws.max_column + 1
            ws.cell(row=header_row, column=col_index, value=col_name)
            col_map[col_name] = col_index

        contrib_dict = df_contrib[col_name].to_dict()
        for r in range(2, ws.max_row + 1):
            cell_date = ws.cell(row=r, column=1).value
            if cell_date is None:
                continue
            try:
                periode = pd.to_datetime(cell_date, errors="coerce").to_period("M")
            except Exception:
                continue
            if periode in contrib_dict:
                ws.cell(row=r, column=col_index, value=float(contrib_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df_contrib, ipc_info



def calculer_contributions_core_noncore_yoy(nom_fichier: str,
                                        feuille_core: str,
                                        feuille_noncore: str,
                                        feuille_categories: str,
                                        date_debut: str,
                                        date_fin: str):
    """
    Calcule la contribution en points de pourcentage (pp) du Core et du Non-Core
    dans l'inflation globale (feuille 'categories').
    Insère les colonnes 'Contrib_Core (pp)' et 'Contrib_Non_Core (pp)' dans la feuille categories.
    Retourne :
        df_contrib : DataFrame avec Contrib_Core et Contrib_Non_Core
        ipc_info   : DataFrame avec IPC_level et IPC_yoy_pct
    """

    # --- 1. Charger les 3 feuilles
    df_core = lire_feuille_wide(nom_fichier, feuille_core)
    df_noncore = lire_feuille_wide(nom_fichier, feuille_noncore)
    df_cat = lire_feuille_wide(nom_fichier, feuille_categories)

    # --- 2. Normaliser les dates
    for df in (df_core, df_noncore, df_cat):
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.to_period("M")
            df.set_index("date", inplace=True)
        else:
            df.index = pd.to_datetime(df.index, errors="coerce").to_period("M")

    # --- 3. Restreindre à la période demandée
    d_debut = pd.Period(date_debut, freq="M")
    d_fin = pd.Period(date_fin, freq="M")
    df_core = df_core.loc[d_debut:d_fin].copy()
    df_noncore = df_noncore.loc[d_debut:d_fin].copy()
    df_cat = df_cat.loc[d_debut:d_fin].copy()

    # --- 4. Charger les poids
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "config" / "weights.json"
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        all_weights = json.load(f)

    poids_core = extraire_poids(all_weights.get(feuille_core, {}))
    poids_noncore = extraire_poids(all_weights.get(feuille_noncore, {}))
    poids_cat = extraire_poids(all_weights.get(feuille_categories, {}))

    # --- 5. Colonnes valides
    colonnes_core = [c for c in df_core.columns if c in poids_core]
    colonnes_noncore = [c for c in df_noncore.columns if c in poids_noncore]
    colonnes_cat = [c for c in df_cat.columns if c in poids_cat]

    if not colonnes_core or not colonnes_noncore or not colonnes_cat:
        raise ValueError("Colonnes manquantes ou incohérence entre Excel et weights.json")

    # --- 6. IPC global
    numer_cat = sum(df_cat[col] * poids_cat[col] for col in colonnes_cat)
    denom_cat = sum(poids_cat[col] for col in colonnes_cat)
    ipc_level = (numer_cat / denom_cat).rename("IPC_level")
    ipc_prev12 = ipc_level.shift(12)

    # --- 7. IPC Core et Non-Core
    numer_core = sum(df_core[col] * poids_core[col] for col in colonnes_core)
    denom_core = sum(poids_core[col] for col in colonnes_core)
    ipc_core = numer_core / denom_core

    numer_noncore = sum(df_noncore[col] * poids_noncore[col] for col in colonnes_noncore)
    denom_noncore = sum(poids_noncore[col] for col in colonnes_noncore)
    ipc_noncore = numer_noncore / denom_noncore

    # --- 8. Contribution Core et Non-Core
    contrib_core = ((ipc_core - ipc_core.shift(12)) / ipc_prev12) * (denom_core / denom_cat) * 100
    contrib_noncore = ((ipc_noncore - ipc_noncore.shift(12)) / ipc_prev12) * (denom_noncore / denom_cat) * 100

    df_contrib = pd.DataFrame({
        "Contrib_Core_YoY (pp)": contrib_core.replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3),
        "Contrib_Non_Core_YoY (pp)": contrib_noncore.replace([np.inf, -np.inf], np.nan).fillna(0.0).round(3),
    })

    # --- 9. IPC global yoy
    ipc_yoy_pct = ((ipc_level - ipc_prev12) / ipc_prev12) * 100
    ipc_info = pd.DataFrame({
        "IPC_level": ipc_level,
        "IPC_yoy_pct": ipc_yoy_pct
    })

    # --- 10. Écriture Excel
    wb = load_workbook(nom_fichier)
    ws = wb[feuille_categories]
    header_row = 1
    col_map = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}

    for col_name in df_contrib.columns:
        if col_name in col_map:
            col_index = col_map[col_name]
        else:
            col_index = ws.max_column + 1
            ws.cell(row=header_row, column=col_index, value=col_name)
            col_map[col_name] = col_index

        contrib_dict = df_contrib[col_name].to_dict()
        for r in range(2, ws.max_row + 1):
            cell_date = ws.cell(row=r, column=1).value
            if cell_date is None:
                continue
            try:
                periode = pd.to_datetime(cell_date, errors="coerce").to_period("M")
            except Exception:
                continue
            if periode in contrib_dict:
                ws.cell(row=r, column=col_index, value=float(contrib_dict[periode]))

    wb.save(nom_fichier)
    wb.close()

    return df_contrib, ipc_info

def pipeline_core_noncore(nom_fichier: str,
                          feuille_core: str,
                          feuille_non_core: str,
                          feuille_categories: str,
                          date_debut: str,
                          date_fin: str):
    """
    Exécute la chaîne complète Core / Non-Core en travaillant sur une copie
    du fichier source ("*_et_calculs.xlsx").
    Étapes :
      1) IPC Core / Non-Core
      2) Inflation MoM Core / Non-Core
      3) Inflation YoY Core / Non-Core
      4) Contributions MoM Core / Non-Core
      5) Contributions YoY Core / Non-Core
    """

    # --- 0. Créer ou utiliser la copie unique ---
    base, ext = os.path.splitext(nom_fichier)
    fichier_calculs = base + "_et_calculs" + ext
    if not os.path.exists(fichier_calculs):
        shutil.copyfile(nom_fichier, fichier_calculs)

    # --- 1. IPC Core / Non-Core ---
    df_ipc_core_noncore = calculer_ipc_core_noncore(
        fichier_calculs, feuille_core, feuille_non_core,
        date_debut, date_fin
    )

    # --- 2. Inflation MoM Core / Non-Core ---
    df_infl_core_mom = calculer_inflation_mom(fichier_calculs, feuille_core, date_debut, date_fin)
    df_infl_noncore_mom = calculer_inflation_mom(fichier_calculs, feuille_non_core, date_debut, date_fin)

    # --- 3. Inflation YoY Core / Non-Core ---
    df_infl_core_yoy = calculer_inflation_yoy(fichier_calculs, feuille_core, date_debut, date_fin)
    df_infl_noncore_yoy = calculer_inflation_yoy(fichier_calculs, feuille_non_core, date_debut, date_fin)

    # --- 4. Contributions MoM Core / Non-Core ---
    df_contrib_core_noncore_mom, ipc_info_mom = calculer_contributions_core_noncore_mom(
        fichier_calculs, feuille_core, feuille_non_core, feuille_categories,
        date_debut, date_fin
    )

    # --- 5. Contributions YoY Core / Non-Core ---
    df_contrib_core_noncore_yoy, ipc_info_yoy = calculer_contributions_core_noncore_yoy(
        fichier_calculs, feuille_core, feuille_non_core, feuille_categories,
        date_debut, date_fin
    )

    return {
        "ipc": df_ipc_core_noncore,
        "infl_core_mom": df_infl_core_mom,
        "infl_noncore_mom": df_infl_noncore_mom,
        "infl_core_yoy": df_infl_core_yoy,
        "infl_noncore_yoy": df_infl_noncore_yoy,
        "contrib_mom": df_contrib_core_noncore_mom,
        "ipc_mom": ipc_info_mom,
        "contrib_yoy": df_contrib_core_noncore_yoy,
        "ipc_yoy": ipc_info_yoy,
    }

def pipeline_calculs(nom_fichier: str,
                     feuille: str,
                     date_debut: str,
                     date_fin: str):
    """
    Exécute la chaîne complète des calculs IPC et inflation
    en travaillant sur une copie unique ("fichier_de_donnes_et_calculs.xlsx").
    """

    # --- 0. Créer ou utiliser la copie unique
    base, ext = os.path.splitext(nom_fichier)
    fichier_calculs = base + "_et_calculs" + ext
    if not os.path.exists(fichier_calculs):
        shutil.copyfile(nom_fichier, fichier_calculs)

    # --- 1. IPC
    df_ipc = calculer_ipc(fichier_calculs, feuille, date_debut, date_fin)

    # --- 2. Inflation éléments MoM
    df_infl_elem_mom = calculer_inflation_elements_mom(fichier_calculs, feuille, date_debut, date_fin)

    # --- 3. Inflation globale MoM
    df_infl_mom = calculer_inflation_mom(fichier_calculs, feuille, date_debut, date_fin)

    # --- 4. Inflation éléments YoY
    df_infl_elem_yoy = calculer_inflation_elements_yoy(fichier_calculs, feuille, date_debut, date_fin)

    # --- 5. Inflation globale YoY
    df_infl_yoy = calculer_inflation_yoy(fichier_calculs, feuille, date_debut, date_fin)

    # --- 6. Contributions MoM
    df_contrib_mom, ipc_info_mom = calculer_contributions_pp_mom(fichier_calculs, feuille, date_debut, date_fin)

    # --- 7. Contributions YoY
    df_contrib_yoy, ipc_info_yoy = calculer_contributions_pp_yoy(fichier_calculs, feuille, date_debut, date_fin)

    return {
        "ipc": df_ipc,
        "infl_elem_mom": df_infl_elem_mom,
        "infl_mom": df_infl_mom,
        "infl_elem_yoy": df_infl_elem_yoy,
        "infl_yoy": df_infl_yoy,
        "contrib_mom": df_contrib_mom,
        "ipc_mom": ipc_info_mom,
        "contrib_yoy": df_contrib_yoy,
        "ipc_yoy": ipc_info_yoy,
    }


import pandas as pd


def get_max_date(nom_fichier: str, feuille: str) -> pd.Timestamp:
    """
    Récupère la date maximale (plus récente) dans l'index d'une feuille Excel.
    """
    df = pd.read_excel(nom_fichier, sheet_name=feuille, index_col=0, parse_dates=True)
    return df.index.max()


def pipeline_global(Fichier_de_donnees: str):
    """
    Fonction globale qui exécute les différents pipelines de calculs
    (Grand Alger, Categories, National, Core/Non-Core).

    Paramètres
    ----------
    Fichier_de_donnees : str
        Chemin vers le fichier Excel contenant toutes les feuilles.
    """

    # --- 1) Dates de référence
    date_debut = "2002-01"  # fixe
    # On va chercher la date max dans chaque feuille
    date_fin_grand_alger = get_max_date(Fichier_de_donnees, "Grand_Alger")
    date_fin_categories = get_max_date(Fichier_de_donnees, "categories")
    date_fin_national = get_max_date(Fichier_de_donnees, "national")
    date_fin_core = get_max_date(Fichier_de_donnees, "core")
    date_fin_non_core = get_max_date(Fichier_de_donnees, "Produits_agricoles_frais")

    # La date de fin globale = la plus récente parmi toutes
    date_fin_globale = max(date_fin_grand_alger,
                           date_fin_categories,
                           date_fin_national,
                           date_fin_core,
                           date_fin_non_core)

    # --- 2) Pipelines individuels
    print("➡️ Pipeline Grand Alger")
    pipeline_calculs(Fichier_de_donnees, "Grand_Alger", date_debut, date_fin_grand_alger.strftime("%Y-%m"))

    print("➡️ Pipeline Categories")
    pipeline_calculs(Fichier_de_donnees, "categories", date_debut, date_fin_categories.strftime("%Y-%m"))

    print("➡️ Pipeline National")
    pipeline_calculs(Fichier_de_donnees, "national", date_debut, date_fin_national.strftime("%Y-%m"))

    # --- 3) Pipeline Core vs Non-Core
    print("➡️ Pipeline Core / Non-Core")
    pipeline_core_noncore(
        nom_fichier=Fichier_de_donnees,
        feuille_core="core",
        feuille_non_core="Produits_agricoles_frais",
        feuille_categories="categories",
        date_debut=date_debut,
        date_fin=date_fin_globale.strftime("%Y-%m")  # on prend la plus récente
    )

    print("✅ Tous les pipelines ont été exécutés avec succès.")


def extraire_inflation_mom(nom_fichier: str, nom_feuille: str, date_ref: str):
    """
    Récupère la valeur de l'inflation mensuelle (Inflation (%, mom))
    à une date donnée (année-mois), ainsi que son évolution par rapport
    au mois précédent.

    Paramètres
    ----------
    nom_fichier : str
        Chemin du fichier Excel
    nom_feuille : str
        Nom de la feuille dans le fichier Excel
    date_ref : str
        Date de référence au format 'YYYY-MM-DD' (seuls année et mois sont pris en compte)

    Retour
    ------
    tuple (str, str)
        - taux_actuel : valeur formatée à la date donnée (ex: '0.85%')
        - evolution : différence vs mois précédent (ex: '+0.23' ou '-0.45')
    """

    # Charger les données
    df = pd.read_excel(nom_fichier, sheet_name=nom_feuille, index_col=0, parse_dates=True)

    col_inflation = "Inflation (%, mom)"
    if col_inflation not in df.columns:
        raise ValueError(f"Colonne '{col_inflation}' introuvable dans {nom_feuille}")

    # Extraire année et mois
    date_ref_dt = pd.to_datetime(date_ref)
    annee, mois = date_ref_dt.year, date_ref_dt.month

    # Filtrer la ligne correspondante (mois demandé)
    mask = (df.index.year == annee) & (df.index.month == mois)
    if not mask.any():
        raise ValueError(f"Aucune donnée pour {annee}-{mois:02d} dans {nom_feuille}")

    taux_actuel = df.loc[mask, col_inflation].iloc[0]

    # Déterminer mois précédent
    if mois == 1:  # si janvier -> comparer à décembre de l'année précédente
        annee_prec, mois_prec = annee - 1, 12
    else:
        annee_prec, mois_prec = annee, mois - 1

    # Récupérer la valeur du mois précédent
    mask_prec = (df.index.year == annee_prec) & (df.index.month == mois_prec)
    if not mask_prec.any():
        raise ValueError(f"Aucune donnée pour {annee_prec}-{mois_prec:02d} (comparaison)")

    taux_precedent = df.loc[mask_prec, col_inflation].iloc[0]

    # Calcul évolution
    evolution = taux_actuel - taux_precedent

    # Formatage
    taux_actuel_fmt = f"{taux_actuel:.2f}%"
    evolution_fmt = f"{evolution:+.2f}"

    return taux_actuel_fmt, evolution_fmt

def extraire_inflation_yoy(nom_fichier: str, nom_feuille: str, date_ref: str):
    """
    Récupère la valeur de l'inflation annuelle (Inflation (%, yoy))
    à une date donnée (année-mois), ainsi que son évolution par rapport
    au même mois de l'année précédente.

    Paramètres
    ----------
    nom_fichier : str
        Chemin du fichier Excel
    nom_feuille : str
        Nom de la feuille dans le fichier Excel
    date_ref : str
        Date de référence au format 'YYYY-MM-DD' (seuls année et mois sont pris en compte)

    Retour
    ------
    tuple (str, str)
        - taux_actuel : valeur formatée à la date donnée (ex: '7.85%')
        - evolution : différence vs même mois année précédente (ex: '+0.23' ou '-0.45')
    """

    # Charger les données
    df = pd.read_excel(nom_fichier, sheet_name=nom_feuille, index_col=0, parse_dates=True)

    col_inflation = "Inflation (%, yoy)"
    if col_inflation not in df.columns:
        raise ValueError(f"Colonne '{col_inflation}' introuvable dans {nom_feuille}")

    # Conversion de la date de référence
    date_ref_dt = pd.to_datetime(date_ref)
    annee, mois = date_ref_dt.year, date_ref_dt.month

    # Filtrer la ligne correspondante (même année-mois)
    mask = (df.index.year == annee) & (df.index.month == mois)
    if not mask.any():
        raise ValueError(f"Aucune donnée pour {annee}-{mois:02d} dans {nom_feuille}")

    taux_actuel = df.loc[mask, col_inflation].iloc[0]

    # Date de comparaison : même mois année précédente
    annee_prec = annee - 1
    mask_prec = (df.index.year == annee_prec) & (df.index.month == mois)
    if not mask_prec.any():
        raise ValueError(f"Aucune donnée pour {annee_prec}-{mois:02d} (comparaison)")

    taux_precedent = df.loc[mask_prec, col_inflation].iloc[0]

    # Calcul évolution
    evolution = taux_actuel - taux_precedent

    # Formatage
    taux_actuel_fmt = f"{taux_actuel:.2f}%"
    evolution_fmt = f"{evolution:+.2f}"

    return taux_actuel_fmt, evolution_fmt



# --- Exemple d'utilisation ---
if __name__ == "__main__":
    Fichier_de_donnes = "Fichier_de_donnes.xlsx"
    pipeline_global(Fichier_de_donnes)






